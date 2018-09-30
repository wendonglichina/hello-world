# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import random
import time
class excel:
    def _init_(self):
        self.a = ""
    def openexcel(self):
        wb = load_workbook("list.xlsx")
        ws = wb["Sheet1"]#wb.active
        list = []
        i = 4
        num = ws["D" + str(i)].value
        while num != None:
            list.append(num)
            i = i + 1
            num = ws["D" + str(i)].value
        return list
    def calcnumber(self,numberlist):
        decNum = {}
        lastN = {}
        for num in numberlist:
            numL = num.split(" ")
            last = numL[-1]
            numLs = numL.pop()
            for n in numL:
                if n not in decNum.keys():
                    decNum[n] = 1
                else:
                    decNum[n] = decNum[n] + 1
            if last not in lastN.keys():
                lastN[last] = 1
            else:
                lastN[last] = lastN[last] + 1
        a = sorted(decNum.items(),key = lambda x:x[1]*random.randint(0,33))
        b = sorted(lastN.items(),key = lambda x:x[1]*random.randint(0,16))
        c = sorted(a[0:6],key = lambda x:x[0])
        new = c[0][0] + " " + c[1][0] + " " + c[2][0] + " " + c[3][0] + " " + c[4][0] + " " + c[5][0] + " " + b[0][0]
        print (new)
        #print("<---------------------------------------------->")
        #print (str(c[0][1]) + " " + str(c[1][1]) + " " + str(c[2][1]) + " " + str(c[3][1]) + " " + str(c[4][1]) + " " + str(c[5][1]) + " " + str(b[0][1]))
        numberlist.append(new)
        return numberlist
if __name__=="__main__":
    e = excel()
    number = e.openexcel()
    #newnumber = e.calcnumber(number)
    while 1:
        
        number = e.calcnumber(number)
        time.sleep(24*3600)
